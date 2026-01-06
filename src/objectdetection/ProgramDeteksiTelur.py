import os
import cv2
import supervision as sv
from ultralytics import YOLO
import pandas as pd
from datetime import datetime
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import time
import urllib.request
import xlwings as xw
import requests
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

os.environ['TF_CPP_MIN_LOG_LEVEL'] = '1'
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'

excelpath = "d:/Kuliah/TUGAS AKHIR/KODINGAN/PYTHON/Excel/DataPengujianIOTI.xlsx"
os.makedirs(os.path.dirname(excelpath), exist_ok=True)

yolomodelocr = YOLO("yolov8-deteksilabelbaru.pt")
yolomodelegg = YOLO("yolov8s-DeteksiTelur.pt")

#source_path = 0 #Untuk webcam ataupun kamera eksternal
source_path = "D:/Kuliah/TUGAS AKHIR/KODINGAN/PYTHON/Data Uji/DATAPENGUJIAN/UJI11.mp4" # Untuk video dari file lokal
#source_path = 'http://192.168.95.161/stream' # Untuk ESP32-CAM

selangwaktu = 0.5
def get_current_time(): 
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def upload_excel_to_drive(filepath, folder_id):
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload

    creds = service_account.Credentials.from_service_account_file(
        'drive_service_account.json',
        scopes=['https://www.googleapis.com/auth/drive']
    )
    service = build('drive', 'v3', credentials=creds)

    filename = os.path.basename(filepath)

    query = f"name = '{filename}' and '{folder_id}' in parents and trashed = false"
    response = service.files().list(q=query, spaces='drive', fields='files(id, name)').execute()
    files = response.get('files', [])

    media = MediaFileUpload(filepath, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    if files:
        file_id = files[0]['id']
        service.files().update(fileId=file_id, media_body=media).execute()
        print(f"♻️ File lama di-drive berhasil di-overwrite: {filename}")
        return f"https://drive.google.com/file/d/{file_id}/view"
    else:
        file_metadata = {'name': filename, 'parents': [folder_id]}
        file = service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink').execute()
        print(f"✅ File baru berhasil diupload ke Drive: {file['webViewLink']}")
        return file['webViewLink']

def sendtoblynk():
    BLYNK_AUTH = '1OwNej3zHj0zCn1eje-H_ZuGH0UKe8OM'
    BASE_URL = f'https://blynk.cloud/external/api/update?token={BLYNK_AUTH}'

    app = xw.App(visible=False)
    app.display_alerts = False
    app.screen_updating = False

    wb = app.books.open(excelpath)
    sheet = wb.sheets.active
    sheet = wb.sheets['Kesimpulan']

    dataproduktivitas1 = sheet.range("B2").value  
    dataproduktivitas2 = sheet.range("B6").value  
    dataproduktivitas3 = sheet.range("B7").value  
    dataproduktivitas4 = sheet.range("B8").value  
    dataproduktivitas5 = sheet.range("B9").value  
    dataproduktivitas6 = sheet.range("B10").value
    dataproduktivitas7 = sheet.range("B11").value
    dataproduktivitas8 = sheet.range("B12").value
    dataproduktivitas9 = sheet.range("B13").value
    dataproduktivitas10 = sheet.range("B3").value
    dataproduktivitas11 = sheet.range("B4").value
    dataproduktivitas12 = sheet.range("B5").value  


    v11 = f'{BASE_URL}&V11={dataproduktivitas1}'  
    v12 = f'{BASE_URL}&V12={dataproduktivitas2}'  
    v13 = f'{BASE_URL}&V13={dataproduktivitas3}'  
    v14 = f'{BASE_URL}&V14={dataproduktivitas4}'  
    v15 = f'{BASE_URL}&V15={dataproduktivitas5}'  
    v16 = f'{BASE_URL}&V16={dataproduktivitas6}' 
    v17 = f'{BASE_URL}&V17={dataproduktivitas7}'  
    v18 = f'{BASE_URL}&V18={dataproduktivitas8}'  
    v19 = f'{BASE_URL}&V19={dataproduktivitas9}'  
    v20 = f'{BASE_URL}&V20={dataproduktivitas10}'
    v21 = f'{BASE_URL}&V21={dataproduktivitas11}'
    v22 = f'{BASE_URL}&V22={dataproduktivitas12}'

    try:
        while True:
            val = requests.get(BASE_URL) 
            vpin11 = requests.get(v11, timeout=5)
            vpin12 = requests.get(v12, timeout=5) 
            vpin13 = requests.get(v13, timeout=5) 
            vpin14 = requests.get(v14, timeout=5) 
            vpin15 = requests.get(v15, timeout=5) 
            vpin16 = requests.get(v16, timeout=5)
            vpin17 = requests.get(v17, timeout=5)
            vpin18 = requests.get(v18, timeout=5)
            vpin19 = requests.get(v19, timeout=5)
            vpin20 = requests.get(v20, timeout=5)
            vpin21 = requests.get(v21, timeout=5)
            vpin22 = requests.get(v22, timeout=5)  
            wb.close()  
            app.quit()  
            print("Data berhasil dikirim ke Blynk.")
            break  
    except requests.exceptions.RequestException as e:
        print(f"Error mengirim data ke Blynk: {e}")

def cek_tidak_produktif(kolom_nilai):
    arr = np.array(kolom_nilai.fillna(-1))
    for i in range(len(arr) - 2):
        if arr[i] == 0 and arr[i+1] == 0 and arr[i+2] == 0:
            return "Tidak Produktif"
    return "Produktif"

def excelwrite(tanggal_list, waktu_list, egg_counts, ocr_list):
    sheet_name = 'Kandang A'
    
    new_data = pd.DataFrame({
        'Tanggal': tanggal_list,
        'Waktu': waktu_list,
        'Egg Count': egg_counts,
        'Ketersediaan Telur': ['ADA' if count > 0 else 'KOSONG' for count in egg_counts],
        'Lokasi': ocr_list
    })

    header_font = Font(bold=True, color="000000", size=12)
    header_fill = PatternFill("solid", fgColor="FFFF00")
    cell_font = Font(bold=True, color="000000", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    red_fill = PatternFill("solid", fgColor="FF0000")
    white_font = Font(bold=True, color="FFFFFF", size=11)

    def format_header(ws, columns):
        for idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=3, column=idx, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        ws.merge_cells('B1:D1')
        ws['B1'] = sheet_name
        ws['B1'].font = Font(bold=True, size=18)
        ws['B1'].alignment = center_align
        ws['B1'].fill = header_fill

        for i in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 15

    def write_data(ws, start_row):
        for i, row in new_data.iterrows():
            for j, value in enumerate(row, start=1):
                cell = ws.cell(row=start_row + i, column=j, value=value)
                cell.font = cell_font
                cell.alignment = center_align

                if j == 1:
                    cell.number_format = 'DD/MM/YYYY'
                elif j == 4 and value == "KOSONG":
                    cell.fill = red_fill
                    cell.font = white_font

    # Tulis data deteksi ke sheet utama
    if os.path.exists(excelpath):
        wb = load_workbook(excelpath)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        last_row = ws.max_row

        if last_row < 3:
            format_header(ws, new_data.columns)
            last_row = 3

        write_data(ws, last_row + 1)
        wb.save(excelpath)
        wb.close()
    else:
        with pd.ExcelWriter(excelpath, engine='openpyxl', mode='w') as writer:
            new_data.to_excel(writer, sheet_name=sheet_name, startrow=2, index=False)
            wb = writer.book
            ws = writer.sheets[sheet_name]

            format_header(ws, new_data.columns)

            for row in range(4, 4 + len(new_data)):
                for col in range(1, len(new_data.columns) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = cell_font
                    cell.alignment = center_align

                    if col == 1:
                        cell.number_format = 'DD/MM/YYYY'
                    elif col == 4 and cell.value == "KOSONG":
                        cell.fill = red_fill
                        cell.font = white_font
            wb.save(excelpath)

    # Sheet Rekap (Pivot)
    df = pd.read_excel(excelpath, sheet_name=sheet_name, header=2)
    df.columns = df.columns.str.strip()
    df = df[['Tanggal', 'Waktu', 'Egg Count', 'Lokasi']].copy()
    df['Tanggal'] = pd.to_datetime(df['Tanggal'], errors='coerce').dt.date
    df['Egg Count'] = pd.to_numeric(df['Egg Count'], errors='coerce')
    df.dropna(subset=['Tanggal', 'Egg Count', 'Lokasi'], inplace=True)

    pivot_df = df.pivot_table(
        index='Tanggal', columns='Lokasi',
        values='Egg Count',
        aggfunc=lambda x: (x > 0).any().astype(int)
    )

    with pd.ExcelWriter(excelpath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        pivot_df.to_excel(writer, sheet_name="Rekap")

    wb = load_workbook(excelpath)
    ws = wb["Rekap"]

    for col_idx in range(1, len(pivot_df.columns) + 2):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 12

    # Sheet Kesimpulan (baru, hanya berisi lokasi dan status produktivitas)
    lokasi_list = sorted(pivot_df.columns.tolist())
    hasil = {loc: cek_tidak_produktif(pivot_df.get(loc, pd.Series([0]*3))) for loc in lokasi_list}

    # Hapus sheet Kesimpulan jika sudah ada, agar tidak duplikat
    if "Kesimpulan" in wb.sheetnames:
        std = wb["Kesimpulan"]
        wb.remove(std)
        wb.save(excelpath)
        wb = load_workbook(excelpath)  # reload

    wss = wb.create_sheet("Kesimpulan")
    wss.append(["Lokasi", "Status Produktivitas"])

    for lokasi in lokasi_list:
        status = hasil.get(lokasi, "Tidak Ada Data")
        wss.append([lokasi, status])

    # Format dan style untuk sheet Kesimpulan
    for i, row in enumerate(wss.iter_rows(min_row=2, max_row=wss.max_row, max_col=2), start=2):
        lokasi_cell, status_cell = row
        lokasi_cell.font = cell_font
        lokasi_cell.alignment = center_align
        status_cell.font = cell_font
        status_cell.alignment = center_align

        if status_cell.value == "Tidak Produktif":
            status_cell.fill = red_fill
            status_cell.font = white_font

    for col_idx in range(1, 3):
        col_letter = get_column_letter(col_idx)
        wss.column_dimensions[col_letter].width = 20

    wb.save(excelpath)
    print("Data, Rekap, dan Kesimpulan berhasil ditulis ke Excel.")

def yolo():
    cap = cv2.VideoCapture(source_path)

    tanggal_list = []
    waktu_list = []
    egg_counts = []
    ocr_list = []

    last_save = datetime.now()
    save_interval = selangwaktu

    while True:
        # Membaca video dari ESP32-CAM
        """ img_resp=urllib.request.urlopen(url)
        imgnp=np.array(bytearray(img_resp.read()),dtype=np.uint8)
        videcap = cv2.imdecode(imgnp,-1) """

        # Membaca video dari file lokal ataupun webcam
        ret, videocap = cap.read()
        if not ret:
            break

        #rotate_image = cv2.rotate(videocap, cv2.ROTATE_90_CLOCKWISE)
        #image_resized = cv2.resize(videocap, (640, 480))
        frame = videocap.copy()

        resultsocr = yolomodelocr(
            source=frame,
            conf=0.65, 
            device=0, 
        )
        resultsegg = yolomodelegg(
            source=frame,
            conf=0.65, 
            device=0, 
        )

        detectionsocr = sv.Detections.from_ultralytics(resultsocr[0])
        detectionsegg = sv.Detections.from_ultralytics(resultsegg[0])

        egg_count = len(detectionsegg.xyxy)
        egg_count = 0
        for i, (box, conf, class_id) in enumerate(zip(detectionsocr.xyxy, detectionsocr.confidence, detectionsocr.class_id)):
                x1, y1, x2, y2 = box
                labelocr = resultsocr[0].names[int(class_id)]

                cv2.rectangle(frame, (int(x1), int(y1)), (int(x2), int(y2)), (0, 255, 0), 2)
                label_size, _ = cv2.getTextSize(labelocr, cv2.FONT_HERSHEY_SIMPLEX, 0.5, 1)
                cv2.rectangle(frame, (int(x1), int(y1) - label_size[1] - 10), 
                            (int(x1) + label_size[0], int(y1)), (0, 255, 0), -1)
                cv2.putText(frame, labelocr, (int(x1), int(y1) - 10), 
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1)
                cv2.putText(frame, f"Nomor Kandang: {labelocr}", (200, 30),
                            cv2.FONT_HERSHEY_DUPLEX, 0.5, (255, 255, 255), 1)
            
                for i, (box, conf, class_id) in enumerate(zip(detectionsegg.xyxy, detectionsegg.confidence, detectionsegg.class_id)):
                        x1, y1, x2, y2 = box
                        labelegg = f"{resultsegg[0].names[int(class_id)]} {conf:.2f}"

                        cv2.rectangle(frame, (int(x1), int(y1)), (int(x2), int(y2)), (0, 255, 0), 2)
                        label_size, _ = cv2.getTextSize(labelegg, cv2.FONT_HERSHEY_SIMPLEX, 0.5, 1)
                        cv2.rectangle(frame, (int(x1), int(y1) - label_size[1] - 10), 
                                            (int(x1) + label_size[0], int(y1)), (0, 255, 0), -1)
                        cv2.putText(frame, labelegg, (int(x1), int(y1) - 10), 
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 1)
                        egg_count = 1

        cv2.putText(frame, f"Telur Terdeteksi: {egg_count}", (10, 30), 
                    cv2.FONT_HERSHEY_DUPLEX, 0.5, (255, 255, 255), 1)
        
        cv2.rectangle(frame, (300, 700), (480, 720), (255, 255, 255), -1)
        cv2.putText(frame, get_current_time(), (300, 710), 
                    cv2.FONT_HERSHEY_DUPLEX, 0.5, (0, 0, 0), 1)

        now = datetime.now()
        if (now - last_save).total_seconds() >= save_interval:
            tanggal_list.append(now.strftime("%Y-%m-%d"))
            waktu_list.append(now.strftime("%H:%M:%S"))
            egg_counts.append(egg_count)
            ocr_list.append(labelocr)
            last_save = now
            print(f"Recorded egg count: {egg_count} at {waktu_list[-1]}")

        cv2.imshow("Egg Detection", frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    excelwrite(tanggal_list, waktu_list, egg_counts, ocr_list)
    folder_id = "1uSAlkonOTXkCjWQ6SLZNawJKt1LwexZn" 

    upload_excel_to_drive(excelpath, folder_id)

    cap.release()
    cv2.destroyAllWindows()

if __name__ == "__main__":
    yolo()
    sendtoblynk()
    print("Program selesai.")
    time.sleep(1)
    print("Proses selesai! Menunggu trigger berikutnya...")
    time.sleep(3)
    exit(0)